# скрипт создания реестра
import datetime
import logging

import openpyxl
import typing

# ключи доступа к файлам словаря данных заявки
from utils.db_api.db_manage import add_record

dict_keys = ['user_id', 'file_name', 'file_id', 'payment_sum', 'payment_amount', 'payment_recipient',
             'purpose_of_payment', 'payment_deadline']

titles = ['Дата поступления', 'Инициатор платежа', 'Основание платежа', 'Сумма платежа', 'Размер оплаты',
          'Получатель', 'Назначение', 'Крайний срок оплаты']


# получаем дату в нужном формате
def get_date() -> str:
    """
    Влзвращает текущую дату и время
    д-м-гггг ч:м:с
    """
    # получаем текущее время
    time_object = datetime.datetime.now()

    # и выделяем день, месяц и год
    today = [str(time_object.day), str(time_object.month), str(time_object.year)]
    # час, минута, секунда
    hms = [str(time_object.hour), str(time_object.minute), str(time_object.second)]
    return '-'.join(today) + ' ' + ':'.join(hms)


def create_book(name: str) -> None:
    """
    Создает реестр в формате excel
    """
    # создание книги в excel
    new_registry = openpyxl.Workbook()
    # удаление таблицы Sheet по-умолчанию
    new_registry.remove(new_registry.active)

    sheet = new_registry.create_sheet(title='Заявки')

    sheet.append(titles)

    new_registry.save(filename=f'registries/{name}')


def create_data_record(data: dict, name: str) -> None:
    """
    Создает новую запись в реестре

    :param data: словарь, который вернет state
    :param name: имя инициатора платежа
    """
    # active_registry = get_active_registry()
    # # открываем книгу с текущей датой создания
    # book = openpyxl.open(filename=f'registries/{active_registry}')
    # sheet = book.worksheets[0]

    # собираем данные в список, устанавляивая типы данных
    data_set = [
        int(data[dict_keys[0]]),
        get_date(),
        name,
        data[dict_keys[1]],
        data[dict_keys[2]],
        int(data[dict_keys[3]]),
        data[dict_keys[4]],
        data[dict_keys[5]],
        data[dict_keys[6]],
        data[dict_keys[7]]
    ]

    add_record(data=data_set)

    # # записываем в строку таблицы
    # sheet.append(data_set)
    # # сохраняем результат
    # book.save(filename=f'registries/{active_registry}')


def get_records_by_name(user_name: str) -> None:
    """
    Заполнит документ user_report по full_name

    :param user_name: ful_name на сервере telegram
    """
    register = openpyxl.open('data/register.xlsx')
    register_sheet = register.worksheets[0]

    report = openpyxl.open('data/user_report.xlsx')
    report.remove_sheet(report.worksheets[0])
    report.create_sheet(title='Заявки')
    report_sheet = report.worksheets[0]

    local_titles = []
    for col in range(1, register_sheet.max_column + 1):
        local_titles.append(register_sheet.cell(row=1, column=col).value)

    report_sheet.append(local_titles)
    for row in range(2, register_sheet.max_row + 1):
        if register_sheet.cell(row=row, column=2).value == user_name:
            report_sheet.append(*register_sheet.iter_rows(min_row=row, max_row=row, values_only=True))
    register.close()
    report.save('data/user_report.xlsx')
